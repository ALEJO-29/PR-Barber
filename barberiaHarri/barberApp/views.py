from django.shortcuts import render
from .models import Cita
from datetime import date, timedelta, datetime
from django.http.response import HttpResponse
from io import BytesIO
from django.http import FileResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from openpyxl import Workbook
from django.views.generic import TemplateView
from django.template.loader import get_template
from xhtml2pdf import pisa


# Create your views here.


def index(request):
    return render(request, ('index.html'))


def reporte(request):
    day = date.today()
    enday = day+timedelta(days=4)
    filtro = Cita.objects.filter(
        fecha__range=[day, enday]).order_by('fecha', 'hora')

    return render(request, 'reporte.html', {'filtro': filtro})


class reporte_excel(TemplateView):
    def get(self, request, *args, **kwargs):
        today = datetime.now()
        query = Cita.objects.filter(fecha=today.date())
        
        wb = Workbook()
        ws = wb.active
        ws['B1'] = "REPORTE DE CITAS"

        ws.merge_cells('B1:E1')

        ws['B3'] = 'FECHA'
        ws['C3'] = 'HORA'
        ws['D3'] = 'TIPO CORTE'
        ws['E3'] = 'NOMBRE'
        ws['F3']= 'APELLID0'
        ws['G3'] = 'TELEFONO'

        cont = 6

        for i in query:
            ws.cell(row=cont, column=2).value = i.fecha
            ws.cell(row=cont, column=3).value = i.hora
            ws.cell(row=cont, column=4).value = i.tipo_corte
            ws.cell(row=cont, column=5).value = i.usuario.nombre
            ws.cell(row=cont, column=6).value = i.usuario.apellido
            ws.cell(row=cont, column=7).value = i.usuario.telefono
            cont += 1
            
        name_file= "ReporteDeCitas.xlsx"
        response= HttpResponse(content_type= "application/ms-excel")
        content= "attachment; filename = {0}".format(name_file)
        response["Content-Disposition"]= content
        
        wb.save(response)
        return response


def pdf(request):
    today = datetime.now()
    query = Cita.objects.filter(fecha=today.date())
    template_path= "pdf.html"
    context= {'query': query}
    
    response= HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte.pdf"'
    
    template= get_template(template_path)
    html= template.render(context)
    
    pisa_status= pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('EXISTE UN ERROR <pre>'+ html+'</pre>')
    return response